import inspect
import logging
from openpyxl import load_workbook
import softest


class Utils(softest.TestCase):
    def assert_list_items_text(self, list, value):
        for stop in list:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def assert_result_registration (self, list, full_name):
        for name in list:
            self.soft_assert(self.assertEqual, name.text, full_name)
            if name.text == full_name:
                pass
        self.assert_all()


    def logging_site(logLevel=logging.DEBUG):
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)

        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("C:\\Users\\user\\PycharmProjects\\Test_Site\\logs\\automation.log", mode="a")
        # create formatter - how you want logs to be formatted
        formatter = logging.Formatter('%(asctime)s -%(filename)s - %(funcName)s - %(levelname)s : %(message)s ', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add console to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist
